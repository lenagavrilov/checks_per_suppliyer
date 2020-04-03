import openpyxl as xl
from openpyxl.styles import Font
import datetime


def get_checks_data(date):
    wb = xl.load_workbook('my_checks.xlsx')
    checks_sheet = wb['checks']
    title = get_title_for_table(wb)
    date_to_compare = datetime.datetime.strptime(date, '%Y-%m-%d')
    all_checks = get_all_checks(checks_sheet)
    updated_checks = get_updated_checks(all_checks, date_to_compare)
    insert_into_excel(updated_checks, title, wb)
    sum_checks_per_supplier()


def get_all_checks(checks_sheet):
    all_checks_list = []

    for row in range(2, checks_sheet.max_row + 1):
        if not row:
            continue
        check_number = checks_sheet.cell(row, 1)
        check_sum = checks_sheet.cell(row, 2)
        check_date = checks_sheet.cell(row, 3)
        check_recipient = checks_sheet.cell(row, 4)
        check_for_what = checks_sheet.cell(row, 5)
        check_giving_date = checks_sheet.cell(row, 6)
        check_comments = checks_sheet.cell(row, 7)
        all_checks_list.append([check_number.value,
                                check_sum.value,
                                check_date.value,
                                check_recipient.value,
                                check_for_what.value,
                                check_giving_date.value,
                                check_comments.value])
    return all_checks_list


def get_updated_checks(all_checks_list, date_to_compare):
    checks_list_updated = []
    for check in all_checks_list:
        if isinstance(check[2], datetime.datetime):
            if check[2] >= date_to_compare:
                check[2] = datetime.datetime.strftime(check[2], '%d/%m/%Y')
                if isinstance(check[5], datetime.datetime):
                    check[5] = datetime.datetime.strftime(check[5], '%d/%m/%Y')
                checks_list_updated.append(check)
    return checks_list_updated


def get_title_for_table(wb):
    sheet = wb['monthly']
    titles = []
    for col in range(2, 9):
        cell = sheet.cell(2, col)
        titles.append(cell.value)
    return titles


def insert_into_excel(checks_list, title, wb):
    for check in checks_list:
        supplier_name = check[3]

        if supplier_name not in wb.sheetnames:
            wb.create_sheet(supplier_name)
            wb[supplier_name].append(title)
            wb[supplier_name].append(check)
        else:
            wb[supplier_name].append(check)
    try:
        wb.save("My_test.xlsx")
    except PermissionError:
        print("The file is open or permission denied.")


def read_relevant_sheetnames_from_created_file():
    wb = xl.load_workbook('My_test.xlsx')
    sheet_names = wb.sheetnames
    list_of_sheetnames = []
    [list_of_sheetnames.append(each) for each in sheet_names
     if sheet_names.index(each) >= 5]
    return list_of_sheetnames


def sum_checks_per_supplier():
    list_of_sheetnames = read_relevant_sheetnames_from_created_file()
    wb = xl.load_workbook('My_test.xlsx')
    for list_name in list_of_sheetnames:
        sheet = wb[list_name]
        rows_quantity = sheet.max_row
        cell_for_sum = sheet.cell(rows_quantity + 1, 2)
        cell_for_sum.font = Font(bold=True)
        sheet.cell(rows_quantity + 1, 1).font = Font(bold=True)
        sheet.cell(rows_quantity + 1, 1).value = 'סה"כ'
        sheet.freeze_panes = "B2"
        # sheet[cell_for_sum] = "=sum(B2:B70)"

        sum = 0
        for row in range(2, rows_quantity + 1):
            cell = sheet.cell(row, 2)
            if isinstance(cell.value, int):
                sum += cell.value
        cell_for_sum.value = sum

    try:
        wb.save("My_test.xlsx")
    except PermissionError:
        print("The file is open or permission denied.")


get_checks_data('2020-03-18')
